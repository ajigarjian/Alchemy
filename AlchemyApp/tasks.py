from celery import shared_task #Added for progress bar functionality
from celery_progress.backend import ProgressRecorder #Added for progress bar functionality
from celery.exceptions import Ignore

import json #for api calls
import re
import random #for generating system color for client dashboard
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse #for API calls
from django.urls import reverse #for HttpResponseRedirect(reverse)
from django.contrib.auth import authenticate, login, logout #for login/logout/register
from django.views.decorators.csrf import csrf_exempt #for API calls
from django.contrib import messages #for register error message(s)
from .models import CustomUser, Client, ControlFamily, System, ControlImplementationStatement #for interacting with database
from .forms import OrganizationForm
from django.contrib.auth.backends import ModelBackend
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required #to redirect user to login route if they try to access an app page past login
from django.db.models import Count, Q, OuterRef, Exists
import os, openai
# from openai import OpenAI
from dotenv import load_dotenv #for access to .env variables, like OPENAI API key
from urllib.parse import unquote #to decode url strings passed through urls as parameters, e.g. client
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from copy import copy
from django.conf import settings
from math import ceil
import xml.etree.ElementTree as ET
import zipfile
import tempfile
import shutil #for removing temporary directories once they are done being used

from langchain.document_loaders import PyPDFLoader #for taking relevant pdfs and loading them as langchain "document" objects
from langchain.document_loaders import Docx2txtLoader #for taking relevant word documents and loading them as langchain "document" objects
from langchain.text_splitter import CharacterTextSplitter #for taking langchain documents and splitting them into chunks (pre-processing for vector storage)
from langchain.embeddings.openai import OpenAIEmbeddings #for taking document chunks and embedding them as vectors for similarity searching
from langchain.vectorstores import FAISS #for storing the vector representations of document chunks, vectorizing the given query, and retrieving the relevant text via similarity search. Will not be long term solution
from langchain.chains import RetrievalQA #Langchain chain for distilling the retrieved document chunks into an human-like answer using an llm/chat model, like gpt-turbo-3.5
from langchain.chains.summarize import load_summarize_chain
from langchain.chains import LLMChain
from langchain.llms import OpenAI
from langchain.chat_models import ChatOpenAI #Importing langchain's abstraction on top of the OpenAI API
from langchain.prompts import PromptTemplate #for creating a prompt template that will provide context to user queries/other inputs to llm

load_dotenv()

@shared_task(bind=True)
def process_assessment_task(self, temp_files):

    try:
        self.update_state(state='STARTED', meta={'current': 0, 'total': 1})

        # Determine the file storage directory based on the environment
        if settings.DEBUG:
            # Development environment
            storage_directory = settings.STATICFILES_DIRS[0]
        else:
            # Production environment
            storage_directory = settings.STATIC_ROOT

        list_of_results = []

        #Store the OpenAI API Key for future use
        openai_api_key = os.getenv("OPENAI_API_KEY", None)

        # Initialize vectorstore and other variables
        vectorstore = None
        relevant_control_families = set()

        # Instantiate the llm we'll be using to tag the files' associated control families
        metadata_llm = OpenAI(model_name="gpt-4", openai_api_key=openai_api_key)

        # Create the prompt template, prompt, and chain we'll use for the tagging
        metadata_template = """Please provide the most relevant FedRAMP control family for a document summarized as follows: {summary}
        Provide nothing but the abbreviation (in all caps) of the control family, e.g. "AT" or "CP" - and nothing else. Your response should only be two letters long."""
        metadata_prompt = PromptTemplate(template=metadata_template, input_variables=["summary"])
        metadata_chain = LLMChain(prompt=metadata_prompt, llm=metadata_llm)

        analyzed_file_summaries = {}
        metadata_llm = OpenAI(model_name='gpt-4', openai_api_key=openai_api_key)

        for index, file_path in enumerate(temp_files): # Loop through all uploaded files

            # Load the temporary file using the PyPDFLoader (if a PDF ) or a DocxLoader (if a Word document)
            if file_path.endswith('.pdf'):
                loader = PyPDFLoader(file_path)
            elif file_path.endswith('.docx'):
                loader = Docx2txtLoader(file_path)
            else:
                return JsonResponse({"error": "Unsupported file type."}, status=400)

            #splitter to chunk up the loaded file. Can tweak chunking parameters for optimal performance 
            text_splitter = CharacterTextSplitter(        
                separator = "\n",
                chunk_size = 1000,
                chunk_overlap  = 200,
                length_function = len,
            )

            # Use the text splitter to chunk up the file into chunks (i.e. "Documents"), each page also has metadata (page number, source file name)
            documents = loader.load_and_split(text_splitter)

            # Run the previously created summarize chain on the Document chunks from the file to get a summary of the file
            file_summary = (load_summarize_chain(metadata_llm, chain_type="stuff")).run(documents)

            #Getting the first few sentences from the file's summary for the dashboard
            sentences = file_summary.split('. ')
            first_two_sentences = '. '.join(sentences[:2]) + ('.' if len(sentences) > 1 else '')
            analyzed_file_summaries[os.path.basename(file_path)] = first_two_sentences

            # Now, with the summary in hand, retrieve the associated control family abbreviation (future state - will be more than just FedRAMP control families)
            file_control_family = metadata_chain.run(file_summary)

            # Tag all the Documents for the given file with the associated control family in its metadata for later use
            for document in documents:
                # Add the 'control_family' field to the metadata
                document.metadata['control_family'] = file_control_family
                # print(document)
                # print("")
            
            relevant_control_families.add(file_control_family)

            if index == 0:
                # Take the first file's document chunks and store them as vector embeddings within FAISS vector storage.
                vectorstore = FAISS.from_documents(documents, OpenAIEmbeddings(openai_api_key=openai_api_key))
            
            else:
                # Add every subsequent file's page chunks into the initialized vector storage
                vectorstore.add_documents(documents)

        # Instantiate the llm we'll be using to analyze the documents using the user's selected OpenAI model
        llm = ChatOpenAI(model_name="gpt-4", temperature=0, openai_api_key=openai_api_key)

        template = """{organization} has implemented security processes based on the provided documents. Please analyze the documents against the following information security test procedure: "{control_description}". None of your responses should contain "I'm sorry", "I apologize" or similar. For this test procedure, please respond with exactly each of the 5 items, with each prefaced by their "Analysis_Part" + the number of the section, e.g. "Analysis_Part1:" (without any introduction or explanation, just your analysis):

        1. Without any introduction or explanation, the snippets of the relevant sections of text that offer evidence to support that the test requirement is implemented. Include the page number, where possible. (If there are no relevant text sections, do not provide anything);
        2. An analysis of how well the document(s) meets the requirements of the given test procedure (If there were no relevant text sections in 1., then explain there was no match);
        3. An implementation status based on 1. and 2. of "Pass" or "Fail" (and nothing else);
        4. If the status was deemed "Fail" in 3., then recommendations for control remediation. If it was deemed "Pass", then do not provide anything; and
        5. Based on the relevant text in 1. and the recommendation in 4., what updated text could look like to meet the procedure. If the status was deemed "Pass" in 3., then do not provide anything.
        """

        # Generating a Langchain prompt template using the string from above
        prompt_template = PromptTemplate.from_template(template)

        org_chain = RetrievalQA.from_chain_type(llm,retriever=vectorstore.as_retriever(), return_source_documents=True)
        document_org = org_chain({"query": "What is the name of the organization that the document is for? Reply with the name and nothing else."})['result']

        #Pull the testing workbook
        static_path = settings.STATIC_ROOT if settings.STATIC_ROOT else settings.STATICFILES_DIRS[0]
        template_path = os.path.join(static_path, 'SRC_Innovation_Documents/FedRAMP Assessment Workbook.xlsx')

        # Load workbook
        workbook = load_workbook(filename=template_path)

        # Create a duplicate of the original workbook for output
        duplicate_workbook = copy(workbook)

        # Before starting the loop over control families, calculate the total number of rows
        total_rows = sum(workbook[family].max_row - 1 for family in relevant_control_families)

        current_progress = 0  # Initialize current progress

        print("pre-processing finished.")

        self.update_state(state='PROGRESS', meta={'current': 0.02, 'total': 1})

        for index, control_family in enumerate(relevant_control_families):
            qa_chain = RetrievalQA.from_chain_type(llm,retriever=vectorstore.as_retriever(search_kwargs={'k': 2, 'filter': {'control_family':control_family}}), return_source_documents=True)

            # select the control family's worksheet in the original workbook and output workbook
            worksheet = workbook[control_family]
            duplicate_worksheet = duplicate_workbook[control_family]

            #Call helper function to actually perform AI-assessment on the procedures for the given control family
            #Pass in all relevant workbook info and AI context as well as the task's self to update the progress within the helper function
            procedure_results, current_progress = assess_controls_in_worksheet(worksheet, duplicate_worksheet, document_org, prompt_template, qa_chain, current_progress, total_rows, self)

            list_of_results.append(procedure_results)
        
        # Save the updated duplicate workbook directly in the static files directory
        workbook_filename = "Assessment_Workbook_with_Analysis.xlsx"
        workbook_file_path = os.path.join(storage_directory, workbook_filename)
        duplicate_workbook.save(workbook_file_path)
        print("Analysis finished. Workbook has been successfully saved at:", workbook_file_path)

        self.update_state(state='PROGRESS', meta={'current': total_rows, 'total': total_rows})

        # -------------------- INTERMEDIARY SECTION TO GET METRICS FROM FILLED OUT WORKBOOK FOR FRONT END ----------------

        # Dictionary to store the intermediate results
        control_intermediate = {}

        # Process each test procedure
        for control_family in list_of_results:
            for procedure, data in control_family.items():
                
                status = data["result"] #Getting each procedure's pass/fail
                if status == 'Partial Pass' or status == 'Error':
                    status = 'Fail'

                control_name = data["name"] #Getting each procedure's 'name' (really control name)
                control = procedure.rsplit('.', 1)[0] #Getting the control ID from each procedure ID

                # If control isn't in the intermediate dictionary yet, initialize it
                if control not in control_intermediate:
                    control_intermediate[control] = {'Pass': 0, 'Fail': 0, 'Name': control_name}

                # Increase the count for the current status (either Pass or Fail) for the current control
                control_intermediate[control][status] += 1

        print("intermediate step complete...")

        # Final dictionary for control results
        control_results = {}

        # Determine the final status for each control
        for control, statuses in control_intermediate.items():
            control_name = statuses.pop('Name')

            if statuses['Pass'] > 0 and statuses['Fail'] == 0:
                status = 'Implemented'
            elif statuses['Pass'] > 0 and statuses['Fail'] > 0:
                status = 'Partially Implemented'
            else:
                status = 'Not Implemented'
            
            control_results[control] = {
                'Name': control_name,
                'Status': status,
                'ProcedurePassCount': statuses['Pass'],
                'ProcedureFailCount': statuses['Fail']
            }
        
        print("control results step complete...")
                
        # --------------------------- METRICS ------------------------------

        # Initialize procedure-specific metric variables
        total_procedures = 0
        passed_procedures_count = 0
        failed_procedures_count = 0

        # Iterate through each control family's results at the procedure level
        for family_results in list_of_results:
            total_procedures += len(family_results)
            passed_procedures_count += sum(1 for data in family_results.values() if data["result"] == 'Pass')

        print("passed procedures complete...")
        
        failed_procedures_count = total_procedures - passed_procedures_count

        #Create control-specific metric variables
        total_controls = len(control_results)
        implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Implemented')
        partially_implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Partially Implemented')
        not_implemented_controls_count = total_controls - implemented_controls_count - partially_implemented_controls_count

        # Pack metrics into a dictionary to send to the frontend

        overview_metrics = {
            'control_results': control_results,
            'total_procedures': total_procedures,
            'passed_procedures_count': passed_procedures_count,
            'failed_procedures_count': failed_procedures_count,
            'total_controls': total_controls,
            'implemented_controls_count': implemented_controls_count,
            'partially_implemented_controls_count': partially_implemented_controls_count,
            'not_implemented_controls_count': not_implemented_controls_count,
            'analyzed_file_summaries': analyzed_file_summaries,
            'number_of_files': len(analyzed_file_summaries)
        }

        print("Metrics creation complete...")

        # ---------------------------------- SECTION TO TEMPORARILY SAVE OUTPUTS (METRICS & TESTING WORKBOOK) FOR FRONT END ---------------------

        # Save the metrics to a JSON file in the static directory
        metrics_filename = "metrics.json"
        metrics_file_path = os.path.join(storage_directory, metrics_filename)
        with open(metrics_file_path, 'w') as json_file:
            json.dump(overview_metrics, json_file)
        print("metrics.json has been successfully created at:", metrics_file_path)
        
        #COMMENTED OUT FOR NOW (RETURNING ZIP VS INDIVIDUAL)
                            # # Move assessment workbook file in temporary path to  static files directory and set up the future zip path
                            # storage_directory = settings.STATICFILES_DIRS[0] if settings.DEBUG else settings.STATIC_ROOT
                            # zip_filename = "Assessment_and_Metrics.zip"
                            # zip_file_path = os.path.join(storage_directory, zip_filename)

                            # #Zip the assessment workbook file and the metrics json file together using the above path
                            # with zipfile.ZipFile(zip_file_path, 'w') as zipf:
                            #     zipf.write(updated_workbook_path, arcname=os.path.basename(updated_workbook_path))
                            #     zipf.write(json_file_path, arcname="metrics.json") # Ensure correct arcname for the JSON file
                            
                            # # Check if the zip file was created and log the result
                            # if os.path.exists(zip_file_path):
                            #     print("Zip file has been successfully created at:", zip_file_path)
                            # else:
                            #     print("Failed to create zip file at:", zip_file_path)
                            
                            # return {'zip_file_path': zip_file_path}

        # Return the paths of the workbook and metrics file
        return {
            'workbook_file_path': workbook_file_path,
            'metrics_file_path': metrics_file_path
        }
    
    except Exception as exc:
        # Properly log and raise the exception with its type
        self.update_state(state='FAILURE', meta={'exc_type': type(exc).__name__, 'exc_message': str(exc)})
        raise Ignore()  # This tells Celery to ignore the result of the task and keep the FAILURE state
    
# Helper function to /generate_ai_assessment that takes in the worksheet, name of the organization, and the prompt template, and updates the workbook accordingly with the output of the analysis
def assess_controls_in_worksheet(worksheet, duplicate_worksheet, organization_name, prompt_template, qa_chain, current_progress, total_rows, task):

    procedure_results = {} #Initialize a dictionary to store the results of each procedure's testing

    starting_row = 2 #Start at the second row to exclude the workbook's header row

    # Loop through each row in the testing workbook and populate the prompt template with the relevant variables
    for row_index, row in enumerate(worksheet.iter_rows(min_row=starting_row, values_only=True), start=starting_row): # Assuming your data starts from the second row (skipping header)
        
        #skip the row if it is empty
        # if row[0]:
        if row_index > 20:

            print(f"Testing control procedure {row[2]}")

            # Getting the current control description from this row in the spreadsheet
            control_description = row[4]

            # Querying the llm to perform a similarity search with the templated query against our vector store
            question = prompt_template.format(control_description=control_description, organization=organization_name)
            result = qa_chain({"query": question})

            source_set = set()

            # Iterate through the source documents for this procedure
            for doc in result["source_documents"]:

                # Replace the full path of the Document's source with just the filename
                doc.metadata['source'] = os.path.basename(doc.metadata['source'])

                # Add the source string to the set
                source_set.add(doc.metadata['source'])
            
            # Join all metadata strings into a single string, separated by a delimiter (e.g., newline)
            all_sources = '\n'.join(source_set)

            # Set the value of the cell to the aggregated metadata string
            cell = duplicate_worksheet.cell(row=row_index, column=6, value=all_sources)
            cell.alignment = Alignment(wrap_text=True)

            # Storing the response string into separate variables by section to upload to different columns
            split_output = extract_parts(result['result'])

            #Taking the split output and populating each row's columns F through K with it
            column_index = 7

            for part in split_output:
                # Add the answer to the current column of the current row in the duplicate worksheet
                cell = duplicate_worksheet.cell(row=row_index, column=column_index, value=part)
                # Set wrap_text to True for the cell
                cell.alignment = Alignment(wrap_text=True)

                #If filling in the pass/fail Column I, then center the text in the cell horizontally and vertically
                if column_index == 9:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Move to the next column for the next loop
                column_index += 1

            procedure_id = row[2]
            procedure_name = row[3]
            procedure_result = duplicate_worksheet.cell(row=row_index, column=9).value

            current_progress += 1
            task.update_state(state='PROGRESS', meta={'current': current_progress, 'total': total_rows})

            procedure_results[procedure_id] = {
                "name": procedure_name,
                "result": procedure_result
            }
    
    return procedure_results, current_progress - 1

# Helper function to /generate_ai_assessment that takes the output from the query to OpenAI from the similarity search, and splits it so that we can put the info into the right columns
def extract_parts(input_str):
    # Define the markers that we will use to split the text
    primary_markers = ["Analysis_Part1:", "Analysis_Part2:", "Analysis_Part3:", "Analysis_Part4:", "Analysis_Part5:"]
    secondary_markers = ["1.", "2.", "3.", "4.", "5."]

    error_message = "I'm sorry, but I can't assist with that."
    if error_message == input_str:
        return ("Error", "Error", "Error", "Error", "Error")
    
    # Function to extract parts using a given set of markers
    def extract_with_markers(markers):
        parts = []
        for i in range(len(markers)):
            start_marker = markers[i]
            try:
                end_marker = markers[i + 1]
            except IndexError:
                end_marker = None

            if end_marker:
                start_idx = input_str.find(start_marker) + len(start_marker)
                end_idx = input_str.find(end_marker)
                if start_idx == -1 or end_idx == -1:
                    return None  # Marker not found
                part = input_str[start_idx:end_idx].strip()
            else:
                start_idx = input_str.find(start_marker) + len(start_marker)
                if start_idx == -1:
                    return None  # Marker not found
                part = input_str[start_idx:].strip()
            parts.append(part)
        return parts

    # Try extracting with primary markers
    extracted_parts = extract_with_markers(primary_markers)
    if extracted_parts is not None and len(extracted_parts) == 5:
        return tuple(extracted_parts)

    # Try extracting with secondary markers
    extracted_parts = extract_with_markers(secondary_markers)
    if extracted_parts is not None and len(extracted_parts) == 5:
        return tuple(extracted_parts)

    # Handle case where neither set of markers worked
    return ('Error', 'Error', 'Error', 'Error', 'Error')
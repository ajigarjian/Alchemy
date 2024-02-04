import json #for api calls
import re
import random #for generating system color for client dashboard
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse #for API calls
from django.urls import reverse #for HttpResponseRedirect(reverse)
from django.contrib.auth import authenticate, login, logout #for login/logout/register
from django.views.decorators.csrf import csrf_exempt #for API calls
from django.contrib import messages #for register error message(s)
from .models import CustomUser, Client, ControlFamily, System, ControlImplementationStatement #for interacting with database
from .forms import OrganizationForm
from django.contrib.auth.backends import ModelBackend
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required #to redirect user to login route if they try to access an app page past login
from django.db.models import Count, Q, OuterRef, Exists
import os, openai
# from openai import OpenAI
from dotenv import load_dotenv #for access to .env variables, like OPENAI API key
from urllib.parse import unquote #to decode url strings passed through urls as parameters, e.g. client
from openpyxl.styles import Alignment
from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook
from copy import copy
from django.conf import settings
from math import ceil
import xml.etree.ElementTree as ET
import zipfile
import tempfile
import shutil #for removing temporary directories once they are done being used
from django.core.files.uploadedfile import TemporaryUploadedFile, InMemoryUploadedFile
import warnings


from .tasks import process_assessment_task #For celery task where AI assessment is performed
from celery.result import AsyncResult #for celery
from celery_progress.backend import ProgressRecorder #for celery

import time #for Kings
from selenium import webdriver #for Kings
from selenium.webdriver.common.by import By #for Kings when selecting elements
from selenium.webdriver.chrome.service import Service #for Kings when importing a ChromeDriver
from selenium.webdriver.chrome.options import Options # for Kings for downloading workbooks
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc #For Kings for undetected ChromeDriver
import csv #for Kings manipulating downloaded csv

from google.cloud import storage #for production serving output files

from langchain.document_loaders import PyPDFLoader #for taking relevant pdfs and loading them as langchain "document" objects
from langchain.document_loaders import Docx2txtLoader #for taking relevant word documents and loading them as langchain "document" objects
from langchain.text_splitter import CharacterTextSplitter #for taking langchain documents and splitting them into chunks (pre-processing for vector storage)
from langchain.embeddings.openai import OpenAIEmbeddings #for taking document chunks and embedding them as vectors for similarity searching
from langchain.vectorstores import FAISS #for storing the vector representations of document chunks, vectorizing the given query, and retrieving the relevant text via similarity search. Will not be long term solution
from langchain.chains import RetrievalQA #Langchain chain for distilling the retrieved document chunks into an human-like answer using an llm/chat model, like gpt-turbo-3.5
from langchain.chains.summarize import load_summarize_chain
from langchain.chains import LLMChain
from langchain.llms import OpenAI
from langchain.chat_models import ChatOpenAI #Importing langchain's abstraction on top of the OpenAI API
from langchain.prompts import PromptTemplate #for creating a prompt template that will provide context to user queries/other inputs to llm

load_dotenv()

####################################### Public Application when not logged in ##############################################

# Suppress specific UserWarning related to langchain
warnings.filterwarnings("ignore", category=UserWarning, message="Importing llm_cache from langchain root module is no longer supported.")
warnings.filterwarnings("ignore", category=UserWarning, message="You are trying to use a chat model. This way of initializing it is no longer supported. Instead, please use: `from langchain.chat_models import ChatOpenAI`")

# Route to render the landing page
def index(request):

    if not request.user.is_authenticated:

        return render(request, "public/index.html")
    
    else:
        return HttpResponseRedirect(reverse("alchemy:dashboard"))

def login_view(request):
    # if this is a POST, process the login data and redirect the logged in user to the overview page
    if request.method == "POST":
        # process the data in form.cleaned_data as required
        email = request.POST["email"]
        password = request.POST["password"]

        #attempt to sign user in, return User object if so
        user = authenticate(request, email=email, password=password)

        if user:
            login(request, user)
            return HttpResponseRedirect(reverse("alchemy:dashboard"))
        else:
            context = {
                'error_message': 'Invalid email or password.',
            }

            return render(request, "public/login.html", context)

    # if not a POST request, show the login page
    if not request.user.is_authenticated:
        return render(request, "public/login.html")
    else:
        return HttpResponseRedirect(reverse("alchemy:logout"))

def register_view(request):

    if request.method == 'POST':
        email = request.POST['email']
        phone_number = request.POST['phone_number']
        client = Client.objects.get(id=request.POST['client'])
        password = request.POST['password']

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect('alchemy:register')

        user = CustomUser.objects.create(email=email, phone_number=phone_number, client=client)
        user.set_password(password)
        user.save()

        # Log the user in.
        backend = ModelBackend()
        user.backend = f"{backend.__module__}.{backend.__class__.__name__}"
        login(request, user)
        return redirect('alchemy:index')

    else:
        if request.user.is_authenticated:
            logout(request)
        return render(request, 'public/register.html', {
            "clients": Client.objects.all()
        })

def logout_view(request):

    if request.user.is_authenticated:
        logout(request)
    return HttpResponseRedirect(reverse("alchemy:index"))

def get_latest_file(download_dir, file_extension=".csv"):
    # Get a list of all files with the specified extension
    files = [f for f in os.listdir(download_dir) if f.endswith(file_extension)]

    # Initialize the latest file name and its last modified time
    latest_file = None
    latest_time = 0

    # Iterate through the files to find the most recently modified
    for file in files:
        file_path = os.path.join(download_dir, file)
        modified_time = os.path.getmtime(file_path)
        if modified_time > latest_time:
            latest_time = modified_time
            latest_file = file

    return latest_file

def contact(request):

    if request.user.is_authenticated:
         return HttpResponseRedirect(reverse("alchemy:dashboard"))
    return render(request, "public/contact.html")

def kings(request):

    #render kings.html
    return render(request, "public/kings.html")

@csrf_exempt
def pull_nfl_data(request):

    # Initialize undetected-chromedriver
    options = uc.ChromeOptions()
    download_dir = str(settings.STATIC_ROOT)
    options.add_experimental_option("prefs", {'download.default_directory' : download_dir})
    driver = uc.Chrome(options=options)

    print("Driver created...")

    # Open the login page
    driver.get("https://www.draftkings.com/lobby#/NFL/0/IsStarred")

    login_page = driver.find_element(By.ID, "sign-up-navigate")
    login_page.click()

    print("At login page...")

    # Find the username, password input fields and the login button
    username = driver.find_element(By.ID, "login-username-input")
    password = driver.find_element(By.ID, "login-password-input")
    login_button = driver.find_element(By.ID, "login-submit")

    username_string = os.getenv("KINGS_USERNAME", None)
    password_string = os.getenv("KINGS_PASSWORD", None)

    print("Retrieved credentials...")

    # Input your username and password
    username.send_keys(username_string)
    time.sleep(2)  # Wait for 2 seconds

    password.send_keys(password_string)
    time.sleep(2)  # Wait for 2 seconds

    # Click the login button
    login_button.click()

    print("Logging in...")

    # Wait for the page to load
    time.sleep(15)

    driver.get("https://www.draftkings.com/lineup/#create-lineup")

    print("At lineup page...")

    time.sleep(2)

    nfl_button = driver.find_element(By.CSS_SELECTOR, "input[type='radio'][name='sport'][value='1']")
    nfl_button.click()
    time.sleep(1)

    game_button = driver.find_element(By.CSS_SELECTOR, "input[type='radio'][name='game-variant'][data-game-variant-id='1']")
    game_button.click()
    time.sleep(1)

    contest_button = driver.find_element(By.CSS_SELECTOR, "input[type='radio'][name='contest-start-date'][value='1']")
    contest_button.click()
    time.sleep(1)

    continue_button = driver.find_element(By.CSS_SELECTOR, "a.dk-btn.dk-btn-primary.continue")
    continue_button.click()
    time.sleep(2)
    
    csv_link = driver.find_element(By.CSS_SELECTOR, "a[data-test-id='player-picker-export-to-csv-link']")
    csv_link.click()
    time.sleep(2)

    print("Daily Fantasy information downloaded to CSV...")

    #Get the name of the CSV file downloaded and get the path
    latest_csv_file_name = get_latest_file(download_dir)
    csv_path = download_dir + "/" + latest_csv_file_name

    #Display the name for tracking purposes
    if latest_csv_file_name:
        print(f"Downloaded CSV file: {latest_csv_file_name}")
    else:
        print("No CSV files found in the directory.")

    #Open the CSV file given its path and convert it to an Excel workbook
    with open(csv_path, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        data = list(reader)
    
    #Instantiate an Excel workbook and copy over the data from the CSV file to an Excel workbook
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)
    
    print("Created XLSX file")

    #Delete the old CSV file now that it's been converted to an Excel workbook
    os.remove(csv_path)
    print("File deleted successfully")

    #Add new rows to the Excel workbook
    header_row = 1  # Assuming first row is the header
    salary_col = ws['E']  # Assuming 'Salary' is in column E
    ws.cell(row=header_row, column=ws.max_column + 1, value="TD Odds")
    ws.cell(row=header_row, column=ws.max_column + 1, value="Yards")
    ws.cell(row=header_row, column=ws.max_column + 1, value="Yards/Cost")
    ws.cell(row=header_row, column=ws.max_column + 1, value="Cost/Yards")

    # Save the workbook with the new columns
    modified_file_path = download_dir + "/" + "Output.xlsx"
    wb.save(modified_file_path)

    driver.get("https://sportsbook.draftkings.com/leagues/football/nfl")
    time.sleep(2)

    # Find all links to upcoming NFL games
    game_elements = driver.find_elements(By.CSS_SELECTOR, '.event-cell-link')

    # Get the links and remove duplicates (they are doubled up due to page structure)
    unique_links = set()
    for element in game_elements:
        href = element.get_attribute('href')
        if href:
            unique_links.add(href)
    
    # Now 'unique_links' contains all unique URLs
    for link in unique_links:
        print(link)
     
    unique_links_list = list(unique_links)
    
    for link in unique_links_list:

        print(f"For the {link} game...")
        
        driver.get(link)

        # Find the ul with class 'component-18-side-rail'
        side_rail = driver.find_element(By.CSS_SELECTOR, "div[aria-labelledby='game_category_Odds'] .component-18-side-rail")

        # Find all the li elements with class 'side-rail-name'
        player_elements = side_rail.find_elements(By.CLASS_NAME, "side-rail-name")

        # Extract and print player names, skipping "No Touchdown Scorer"
        player_names = [player.text.strip() for player in player_elements if player.text.strip() != "No Touchdown Scorer"]

        #Get the TD odds for each players
        # Find the div with aria-labelledby='game_category_Odds'
        odds_category_div = driver.find_element(By.CSS_SELECTOR, "div[aria-labelledby='game_category_Odds']")

        # Find all the 'component-18' divs within the 'game_category_Odds' div
        component_18_divs_within_odds_category = odds_category_div.find_elements(By.CLASS_NAME, "component-18")

        # Select the component div with the TD Odds the second or third one

        if len(component_18_divs_within_odds_category) == 3:
            component_18_div = component_18_divs_within_odds_category[2]
        
        else:
            component_18_div = component_18_divs_within_odds_category[1]

        # Find all the span elements with the desired class within this div
        odds_elements = component_18_div.find_elements(By.CSS_SELECTOR, "span.sportsbook-odds.american.no-margin.default-color")

        # Extract the odds values
        td_values = [odds_element.text.strip() for odds_element in odds_elements]

        # Create a dictionary mapping player names to their odds
        player_odds_dict = dict(zip(player_names, td_values))

        # Print the dictionary
        for player, odds in player_odds_dict.items():
            print(f"{player}: {odds}")
        
        players_odds = {}

        # Extract and map player names to their odds
        for player, odds in player_odds_dict.items():
            players_odds[player] = odds
        
        #Update the Excel file with the odds
        for row in ws.iter_rows(min_row=2):  # Assuming first row is header
            player_name_in_excel = row[2].value  # Assuming player names are in column C
            if player_name_in_excel in players_odds:
                # Assuming you want to update column J
                row[9].value = players_odds[player_name_in_excel]  
        
        #LOGIC FOR ADDING PASS YARDS

        # Check if the "Event Accordion for Pass Yards exists
        pass_accordion_exists = len(driver.find_elements(By.XPATH, "//div[@aria-label='Event Accordion for Pass Yards']")) > 0

        if pass_accordion_exists:
            # Locate the parent div using the aria-label attribute for "Event Accordion for Pass Yards"
            event_accordion_div = driver.find_element(By.XPATH, "//div[@aria-label='Event Accordion for Pass Yards']/following-sibling::div")

            # Find the table body within this specific div
            table_body = event_accordion_div.find_element(By.CSS_SELECTOR, "table.sportsbook-table tbody")

            # Initialize a dictionary to store the players and their pass yards
            pass_yards_dict = {}

            # Iterate through each row in the table
            for row in table_body.find_elements(By.TAG_NAME, "tr"):
                # Get the player's name
                player_name = row.find_element(By.CSS_SELECTOR, ".sportsbook-row-name").text.strip()
                
                # Get the player's pass yards value from the span with the pass yard line
                pass_yards_value = row.find_element(By.CSS_SELECTOR, ".sportsbook-outcome-cell__label-line-container span:last-child").text.strip()

                # Add the player's name and pass yards value to the dictionary
                pass_yards_dict[player_name] = pass_yards_value

            # Print the result
            for player, yards in pass_yards_dict.items():
                print(f"{player}: {yards} passing yards")
        
        #LOGIC FOR ADDING RUSHING YARDS
        # Check if the "Event Accordion for Rush Yards" exists
        rush_accordion_exists = len(driver.find_elements(By.XPATH, "//div[@aria-label='Event Accordion for Rush Yards']")) > 0

        if rush_accordion_exists:
            # Locate the parent div using the aria-label attribute for "Event Accordion for Rush Yards"
            event_accordion_div = driver.find_element(By.XPATH, "//div[@aria-label='Event Accordion for Rush Yards']/following-sibling::div")

            # Find the table body within this specific div
            table_body = event_accordion_div.find_element(By.CSS_SELECTOR, "table.sportsbook-table tbody")

            # Initialize a dictionary to store the players and their pass yards
            rush_yards_dict = {}

            # Iterate through each row in the table
            for row in table_body.find_elements(By.TAG_NAME, "tr"):
                # Get the player's name
                player_name = row.find_element(By.CSS_SELECTOR, ".sportsbook-row-name").text.strip()
                
                # Get the player's pass yards value from the span with the pass yard line
                rush_yards_value = row.find_element(By.CSS_SELECTOR, ".sportsbook-outcome-cell__label-line-container span:last-child").text.strip()

                # Add the player's name and pass yards value to the dictionary
                rush_yards_dict[player_name] = rush_yards_value

            # Print the result
            for player, yards in rush_yards_dict.items():
                print(f"{player}: {yards} rushing yards")
            
            #Update the Excel file with the odds
            for row in ws.iter_rows(min_row=2):  # Assuming first row is header
                player_name_in_excel = row[2].value  # Assuming player names are in column C
                if player_name_in_excel in rush_yards_dict:
                    # Assuming you want to update column J
                    row[10].value = rush_yards_dict[player_name_in_excel]  

            wb.save(modified_file_path)

        
        #LOGIC FOR ADDING RECEIVING YARDS

        # Check if the "Event Accordion for Rec Yards" exists
        rec_accordion_exists = len(driver.find_elements(By.XPATH, "//div[@aria-label='Event Accordion for Receiving Yards']")) > 0

        if rec_accordion_exists:
            # Locate the parent div using the aria-label attribute for "Event Accordion for Receiving Yards"
            event_accordion_div = driver.find_element(By.XPATH, "//div[@aria-label='Event Accordion for Receiving Yards']/following-sibling::div")

            # Find the table body within this specific div
            table_body = event_accordion_div.find_element(By.CSS_SELECTOR, "table.sportsbook-table tbody")

            # Initialize a dictionary to store the players and their pass yards
            receiving_yards_dict = {}

            # Iterate through each row in the table
            for row in table_body.find_elements(By.TAG_NAME, "tr"):
                # Get the player's name
                player_name = row.find_element(By.CSS_SELECTOR, ".sportsbook-row-name").text.strip()
                
                # Get the player's pass yards value from the span with the pass yard line
                receiving_yards_value = row.find_element(By.CSS_SELECTOR, ".sportsbook-outcome-cell__label-line-container span:last-child").text.strip()

                # Add the player's name and pass yards value to the dictionary
                receiving_yards_dict[player_name] = receiving_yards_value

            # Print the result
            for player, yards in receiving_yards_dict.items():
                print(f"{player}: {yards} receiving yards")
            
            #Update the Excel file with the odds
            for row in ws.iter_rows(min_row=2):  # Assuming first row is header
                player_name_in_excel = row[2].value  # Assuming player names are in column C
                if player_name_in_excel in receiving_yards_dict:
                    # Assuming you want to update column J
                    row[10].value = receiving_yards_dict[player_name_in_excel]  

            wb.save(modified_file_path)
        
        #LOGIC FOR ADDING RUSH + REC YARDS FOR ELIGIBLE PLAYERS

        # Check if the "Event Accordion for Rush + Rec Yards" exists
        rush_rec_accordion_exists = len(driver.find_elements(By.XPATH, "//div[@aria-label='Event Accordion for Rush + Rec Yards']")) > 0

        if rush_rec_accordion_exists:
            # Locate the parent div using the aria-label attribute for "Event Accordion for Pass Yards"
            event_accordion_div = driver.find_element(By.XPATH, "//div[@aria-label='Event Accordion for Rush + Rec Yards']/following-sibling::div")

            # Find the table body within this specific div
            table_body = event_accordion_div.find_element(By.CSS_SELECTOR, "table.sportsbook-table tbody")

            # Initialize a dictionary to store the players and their pass yards
            rush_rec_yards_dict = {}

            # Iterate through each row in the table
            for row in table_body.find_elements(By.TAG_NAME, "tr"):
                # Get the player's name
                player_name = row.find_element(By.CSS_SELECTOR, ".sportsbook-row-name").text.strip()
                
                # Get the player's pass yards value from the span with the pass yard line
                rush_rec_yards_value = row.find_element(By.CSS_SELECTOR, ".sportsbook-outcome-cell__label-line-container span:last-child").text.strip()

                # Add the player's name and pass yards value to the dictionary
                rush_rec_yards_dict[player_name] = rush_rec_yards_value

            # Print the result
            for player, yards in rush_rec_yards_dict.items():
                print(f"{player}: {yards} rushing + receiving yards")
            
            #Update the Excel file with the odds
            for row in ws.iter_rows(min_row=2):  # Assuming first row is header
                player_name_in_excel = row[2].value  # Assuming player names are in column C
                if player_name_in_excel in rush_rec_yards_dict:
                    # Assuming you want to update column J
                    row[10].value = rush_rec_yards_dict[player_name_in_excel]  

            wb.save(modified_file_path)
    
    # Iterate over the rows in the worksheet
    for row in ws.iter_rows(min_row=2):  # Assuming first row is header
        yards_value = row[10].value  # Column K (11th column, index 10)
        cost_value = row[5].value   # Column E (5th column, index 4)
        print(yards_value)
        print(cost_value)

        # Check if both yards and cost values are present and numeric
        if (yards_value and cost_value):
            # Calculate Yards/Cost and update the cell value
            yards_per_cost = float(yards_value) / float(cost_value)
            row[11].value = yards_per_cost  # Column L (Yards/Cost, index 11)

        else:
            # If either value is missing or not numeric, leave the cell blank
            row[11].value = None
    
    # Save the workbook with the updated data
    wb.save(modified_file_path)

    return HttpResponse("Data processing completed successfully.")

# Helper function to add/update stats for a player
def add_player_stat(stat_dictionary, player_name, stat, value):
    if player_name not in stat_dictionary:
        # If the player is not already in the dictionary, add them with an empty stats dictionary
        stat_dictionary[player_name] = {}
    # Update the player's stats dictionary with the new stat
    stat_dictionary[player_name][stat] = value

@csrf_exempt
def pull_nba_data(request):

    # Initialize undetected-chromedriver
    options = uc.ChromeOptions()
    download_dir = str(settings.STATIC_ROOT)
    options.add_experimental_option("prefs", {'download.default_directory' : download_dir})
    driver = uc.Chrome(options=options)

    wait = WebDriverWait(driver, 3)  # Delay to make sure page elements load

    print("Driver created...")

    #Go to the Daily Fantasy Fuel website and get the most recent CSV
    try:
        driver.get("https://www.dailyfantasyfuel.com/nba/projections/")
        csv_link = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(@class,'projections-download') and contains(.,'Download CSV')]")))
        csv_link.click()
    except Exception as e:
        print(f"Error finding or clicking the CSV download link: {e}")
    
    #Given the CSV file, port it over to an XLSX file and add the additional columns
    try:
        #Get the name of the CSV file downloaded and get the path
        latest_csv_file_name = get_latest_file(download_dir)
        csv_path = download_dir + "/" + latest_csv_file_name

        #Display the name for tracking purposes
        if latest_csv_file_name:
            print(f"Downloaded CSV file: {latest_csv_file_name}")
        else:
            print("No CSV files found in the directory.")

        #Open the CSV file given its path and convert it to an Excel workbook
        with open(csv_path, mode='r', encoding='utf-8') as file:
            reader = csv.reader(file)
            data = list(reader)
        
        #Instantiate an Excel workbook and copy over the data from the CSV file to an Excel workbook
        wb = Workbook()
        ws = wb.active

        for row in data:
            ws.append(row)
        
        # modified_file_path = download_dir + "/" + "NBA_Report.xlsx"
        # wb.save(modified_file_path)
        
        print("Created .xlsx file")

        #Delete the old CSV file now that it's been converted to an Excel workbook
        os.remove(csv_path)
        print("File deleted successfully")

        #Add new columns to the Excel workbook for player's individual projected stats
        header_row = 1  # Assuming first row is the header
        salary_col = ws['M']  # Assuming 'Salary' is in column M

        ws.cell(row=header_row, column=ws.max_column + 1, value="Points")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Threes")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Rebounds")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Assists")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Steals")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Blocks")

        ws.cell(row=header_row, column=ws.max_column + 1, value="Projected Points")
        ws.cell(row=header_row, column=ws.max_column + 1, value="Cost Per Point")

        # Assuming your data starts from row 2 to avoid headers
        for row in range(2, ws.max_row + 1):
            # Set Column AB (28) to calculate projected points once those rows are populated with individual stats
            ws.cell(row=row, column=28).value = f"=(V{row}*1) + (W{row}*.5) + (X{row}*1.2) + (Y{row}*1.5) + (Z{row}*3) + (AA{row}*3)"
            ws.cell(row=row, column=29).value = f"=(M{row})/(AB{row})"

    except Exception as e:
        print(f"Error saving and modifying the Excel workbook: {e}")
    
    print("Excel workbook formatted")
    
    # Setting the column number that the stats columns are in the worksheet
    stats_columns = { "Points": 22, "Threes": 23, "Rebounds": 24, "Assists": 25, "Steals": 26, "Blocks": 27 }
    
    driver.get("https://sportsbook.draftkings.com/leagues/basketball/nba")

    # Find all links to upcoming NFL games
    game_elements = driver.find_elements(By.CSS_SELECTOR, '.event-cell-link')

    # Get the links and remove duplicates (they are doubled up due to page structure)
    unique_links = set()
    for element in game_elements:
        href = element.get_attribute('href')
        if href:
            unique_links.add(href)
    
    unique_links_list = list(unique_links)

    print(unique_links_list)
    
    for link in unique_links_list:

        print(f"For the {link} game...")

        # Initialize a dictionary to hold all player data
        players_stats = {}

        try:
            driver.get(link)

            offensive_stats = ["Points", "Threes", "Rebounds", "Assists"]
            defensive_stats = ["Steals", "Blocks"]

            time.sleep(2)  # Wait for 2 seconds to let things load

            for stat in offensive_stats:

                # Find the div with the specific aria-label for Points
                stats_div = driver.find_element(By.XPATH, f'//div[@aria-label="Event Accordion for {stat}"]')

                # Move to the adjacent div that contains the table
                table_wrapper_div = stats_div.find_element(By.XPATH, './following-sibling::div')

                # Now find the table within this div
                table = table_wrapper_div.find_element(By.TAG_NAME, 'tbody')

                #get points
                player_stat_values = {}

                # Iterate through each row in the table body to extract player names and "over" values
                for row in table.find_elements(By.TAG_NAME, 'tr'):
                    # Extract the player name, which is in a span with class 'sportsbook-row-name'
                    player_name = row.find_element(By.CSS_SELECTOR, 'span.sportsbook-row-name').text

                    # Extract the "over" value, which is the first td element in the row
                    stat_parent_element = row.find_elements(By.TAG_NAME, 'td')[0]

                    stat_value = stat_parent_element.find_element(By.CSS_SELECTOR, 'span.sportsbook-outcome-cell__line').text

                    # Add the player name and "over" value to the dictionary
                    player_stat_values[player_name] = stat_value

                # Add offensive stat data
                for player_name, value in player_stat_values.items():
                    add_player_stat(players_stats, player_name, stat, value)
            
            defense_tab_link = driver.find_element(By.ID, "subcategory_Player Defense").get_attribute('href')
            driver.get(defense_tab_link)

            for stat in defensive_stats:

                # Find the div with the specific aria-label for Points
                stats_div = driver.find_element(By.XPATH, f'//div[@aria-label="Event Accordion for {stat}"]')

                # Move to the adjacent div that contains the table
                table_wrapper_div = stats_div.find_element(By.XPATH, './following-sibling::div')

                # Now find the table within this div
                table = table_wrapper_div.find_element(By.TAG_NAME, 'tbody')

                #get points
                player_stat_values = {}

                # Iterate through each row in the table body to extract player names and "over" values
                for row in table.find_elements(By.TAG_NAME, 'tr'):
                    # Extract the player name, which is in a span with class 'sportsbook-row-name'
                    player_name = row.find_element(By.CSS_SELECTOR, 'span.sportsbook-row-name').text

                    # Extract the "over" value, which is the first td element in the row
                    stat_parent_element = row.find_elements(By.TAG_NAME, 'td')[0]

                    stat_value = stat_parent_element.find_element(By.CSS_SELECTOR, 'span.sportsbook-outcome-cell__line').text

                    # Add the player name and "over" value to the dictionary
                    player_stat_values[player_name] = stat_value

                # Add offensive stat data
                for player_name, value in player_stat_values.items():
                    add_player_stat(players_stats, player_name, stat, value)

            print(players_stats)

            # Now that we have the players' stats for the given game, we add them to the workbook
            for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
                # Combine the player's first name and last name since they're split in the first two columns
                player_name = f"{row[0].value} {row[1].value}"
                
                # Check if this player is in the dictionary
                if player_name in players_stats:
                    # Get the player's stats
                    player_stats = players_stats[player_name]
                
                    # Iterate through each stat and fill in the values
                    for stat, col in stats_columns.items():
                        if stat in player_stats:
                            ws.cell(row=row[0].row, column=col, value=player_stats[stat])

        except Exception as e:
            print(f"Error pulling the NBA stats from for the {link} game: {e}")
    
    print("NBA Stats inserted into workbook from DK.")
        
    # Once all modifications are done, convert the workbook to a byte stream
    virtual_workbook = save_virtual_workbook(wb)

    # Prepare the HttpResponse to send the Excel file to the client
    response = HttpResponse(virtual_workbook, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="NBA Report.xlsx"'

    driver.close()
    
    return response

def assess(request):

    # render assess.html
    return render(request, "public/assess.html")

def save_to_cloud_storage(bucket_name, source_file_path, destination_blob_name):
    """Uploads a file to Google Cloud Storage."""
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)

    blob.upload_from_filename(source_file_path)
    blob.make_public()

@csrf_exempt
def generate_ai_assessment(request):

    try:

        # Cannot call this API call to OpenAI unless via the file upload POSTing to backend
        if request.method != "POST":
            return JsonResponse({"error": "POST request required."}, status=400)

        # Store the OpenAI API Key for future use
        openai_api_key = os.getenv("OPENAI_API_KEY", None)

        #Figure out if we're in testing or not 
        use_celery = os.getenv("USE_CELERY", None)
        # Convert the string values to booleans
        use_celery = True if use_celery.lower() == "true" else False
        
        # Get the relevant info from the user (files, settings)
        uploaded_files = request.FILES.getlist('data_files')

        print(f"0. Scanning uploaded files {uploaded_files}...")
        
        selected_framework = request.POST['selectedFramework']
        # selected_family = request.POST['selectedFamily']
        selected_model = request.POST['selectedModel']

        temp_files = []

        # for uploaded_file in uploaded_files:
        #     temp_dir = tempfile.mkdtemp()
        #     temp_file_path = os.path.join(temp_dir, uploaded_file.name)
        #     with open(temp_file_path, 'wb') as temp_file:
        #         for chunk in uploaded_file.chunks():
        #             temp_file.write(chunk)
            
        #     temp_files.append(temp_file_path)

        for uploaded_file in uploaded_files:

            if isinstance(uploaded_file, TemporaryUploadedFile):
                # print("Temporary file detected...")
                # Use the file's temporary path directly
                temp_files.append(uploaded_file.temporary_file_path())
            else:
                # For InMemoryUploadedFile, continue to write to a temp file
                temp_dir = tempfile.mkdtemp()
                temp_file_path = os.path.join(temp_dir, uploaded_file.name)
                with open(temp_file_path, 'wb') as temp_file:
                    for chunk in uploaded_file.chunks():
                        temp_file.write(chunk)
                temp_files.append(temp_file_path)

        # ------------ MAIN PROCESSING AREA ------------------

        if use_celery: #If we want to toggle using the celery, rabbitmq, redis process logic, use this branch
            # Call the Celery task and wait for its result
            task_result = process_assessment_task.delay(temp_files)
            task_id = task_result.id

            print(f"1. Asynchronously processing task: {task_id}...")

            # Return the task ID to the frontend
            return JsonResponse({'task_id': task_id})
        
        else:
            # Synchronous processing logic without using celery, rabbitmq, redis, etc.
            print("1. Synchronously processing analysis...")
            result = process_assessment(temp_files)  # Implement this function
            print("Assessment results received.")
            return JsonResponse(result)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def process_assessment(temp_files):

    list_of_results = []

    #Store the OpenAI API Key for future use
    openai_api_key = os.getenv("OPENAI_API_KEY", None)

    # Initialize vectorstore and other variables
    vectorstore = None
    relevant_control_families = set()

    # Instantiate the llm we'll be using to tag the files' associated control families and the files' types (policy, plan, etc.)
    metadata_llm = OpenAI(model_name="gpt-4", openai_api_key=openai_api_key)

    # Create the prompt template, prompt, and chain we'll use for the tagging
    family_template = """Please provide the most relevant FedRAMP control family for a document summarized as follows: {summary}
    Provide nothing but the abbreviation (in all caps) of the control family, e.g. "AT" or "CP" - and nothing else. Your response should only be two letters long."""
    family_prompt = PromptTemplate(template=family_template, input_variables=["summary"])
    family_chain = LLMChain(prompt=family_prompt, llm=metadata_llm)

    analyzed_file_summaries = {}

    file_type_template = """Please provide the type of document for a document with a file name of '{file_name}' summarized as follows: {summary}
    Provide nothing but one of these options for the type of document: "policy", "procedure", "plan", "diagram", "contract", or "report".
    Please provide one of those options and nothing else. Your response should only be one word long."""
    file_type_prompt = PromptTemplate(template=file_type_template, input_variables=["file_name", "summary"])
    file_type_chain = LLMChain(prompt=file_type_prompt, llm=metadata_llm)

    print("2. Beginning pre-processing...")

    summary_chain = (load_summarize_chain(metadata_llm, chain_type="stuff"))

    for index, file_path in enumerate(temp_files): # Loop through all uploaded files
        try:
            # Load the temporary file using the PyPDFLoader (if a PDF ) or a DocxLoader (if a Word document)
            if file_path.endswith('.pdf'):
                loader = PyPDFLoader(file_path)
            elif file_path.endswith('.docx'):
                loader = Docx2txtLoader(file_path)
            else:
                return JsonResponse({"error": "Unsupported file type."}, status=400)

            #splitter to chunk up the loaded file. Can tweak chunking parameters for optimal performance 
            text_splitter = CharacterTextSplitter(        
                separator = "\n",
                chunk_size = 1000,
                chunk_overlap  = 200,
                length_function = len,
            )

            # Use the text splitter to chunk up the file into chunks (i.e. "Documents"), each page also has metadata (page number, source file name)
            documents = loader.load_and_split(text_splitter)

            print("3. File split up, creating metadata such as relevant control families and summary...")

            try:
                # Run the previously created summarize chain on the Document chunks from the file to get a summary of the file
                file_summary = summary_chain.run(documents[:10])
            except:
                print('An error occurred for the file summary generation.')

            #Getting the first few sentences from the file's summary for the dashboard
            sentences = file_summary.split('. ')
            first_two_sentences = '. '.join(sentences[:2]) + ('.' if len(sentences) > 1 else '')
            # print(f"{os.path.basename(file_path)}: {first_two_sentences}")
            analyzed_file_summaries[os.path.basename(file_path)] = first_two_sentences

            # Now, with the summary in hand, retrieve the associated control family abbreviation (future state - will be more than just FedRAMP control families)
            file_control_family = family_chain.run(file_summary)

            file_type = file_type_chain.run(file_name=(os.path.basename(file_path)), summary=file_summary)

            print(f'4. {os.path.basename(file_path)}: {file_type} covering the {file_control_family} family. {first_two_sentences}.')

            # Tag all the Documents for the given file with the associated control family in its metadata for later use
            for document in documents:
                # Add the 'control_family' field to the metadata
                document.metadata['control_family'] = file_control_family
                document.metadata['file_type'] = file_type

            
            relevant_control_families.add(file_control_family)

            print(document)

            if vectorstore == None:
                # Take the first file's document chunks and store them as vector embeddings within FAISS vector storage.
                vectorstore = FAISS.from_documents(documents, OpenAIEmbeddings(openai_api_key=openai_api_key))
            
            else:
                # Add every subsequent file's page chunks into the initialized vector storage
                vectorstore.add_documents(documents)
        except:
            print(f"Error while vectorizing the file at {file_path}...")

    # Instantiate the llm we'll be using to analyze the documents using the user's selected OpenAI model
    llm = ChatOpenAI(model_name="gpt-4", temperature=0, openai_api_key=openai_api_key)

    template = """{organization} has implemented security processes based on the provided documents. They have hired us as security advisors. Please analyze the documents against the following information security test procedure: "{control_description}". None of your responses should contain "I'm sorry", "I apologize" or similar. For this test procedure, please respond with exactly each of the 5 items, with each prefaced by their "Analysis_Part" + the number of the section, e.g. "Analysis_Part1:" (without any introduction or explanation, just your analysis):

    1. Without any introduction or explanation, the snippets of the relevant sections of text that offer evidence to support that the test requirement is implemented. Include the page number, where possible. (If there are no relevant text sections, do not provide anything);
    2. An analysis of how well the document(s) meets the requirements of the given test procedure (If there were no relevant text sections in 1., then explain there was no match);
    3. An implementation status based on 1. and 2. of "Pass" or "Fail" (and nothing else);
    4. If the status was deemed "Fail" in 3., then recommendations for control remediation. If it was deemed "Pass", then do not provide anything; and
    5. Based on the relevant text in 1. and the recommendation in 4., what updated text could look like to meet the procedure. If the status was deemed "Pass" in 3., then do not provide anything.

    Note: If you think that there is no context or documents to analyze in the first place, mention in 1. and 2. that there was no relevant information, make 3. "Fail", and write 4. and 5. as if wholly new text must be added.
    """

    # Generating a Langchain prompt template using the string from above
    prompt_template = PromptTemplate.from_template(template)

    org_chain = RetrievalQA.from_chain_type(llm,retriever=vectorstore.as_retriever(), return_source_documents=True)
    document_org = org_chain({"query": "What is the name of the organization that the document is for? Reply with the name and nothing else."})['result']

    #Pull the testing workbook
    static_path = settings.STATIC_ROOT if settings.STATIC_ROOT else settings.STATICFILES_DIRS[0]
    template_path = os.path.join(static_path, 'SRC_Innovation_Documents/FedRAMP Assessment Workbook.xlsx')

    # Load workbook
    workbook = load_workbook(filename=template_path)

    # Create a duplicate of the original workbook for output
    duplicate_workbook = copy(workbook)

    print("5. Pre-processing finished. Performing analysis...")

    for index, control_family in enumerate(relevant_control_families):

        try:
            # Gather unique sources for the current control family
            unique_sources = set(doc.metadata['source'] for doc in vectorstore.documents if doc.metadata['control_family'] == control_family)

            print(unique_sources)
        except:
            print("Couldn't get sources...")

        qa_chain = RetrievalQA.from_chain_type(llm,retriever=vectorstore.as_retriever(search_kwargs={'k': 2, 'filter': {'control_family':control_family}}), return_source_documents=True)

        # select the control family's worksheet in the original workbook and output workbook
        worksheet = workbook[control_family]
        duplicate_worksheet = duplicate_workbook[control_family]

        #Call helper function to actually perform AI-assessment on the procedures for the given control family
        procedure_results = assess_controls_in_worksheet(worksheet, duplicate_worksheet, document_org, prompt_template, qa_chain)

        print(f"5x. Received procedure results: {procedure_results} for {control_family} family...")

        list_of_results.append(procedure_results)
    
    # Save the updated duplicate workbook directly in the static files directory
    workbook_filename = "Assessment_Workbook_with_Analysis.xlsx"

    if settings.DEBUG == True:
        # Local storage
        workbook_file_path = os.path.join(settings.STATICFILES_DIRS[0], workbook_filename)
        duplicate_workbook.save(workbook_file_path)

    else:
        # Cloud storage
        temp_workbook_path = os.path.join(tempfile.mkdtemp(), workbook_filename)
        duplicate_workbook.save(temp_workbook_path)
        save_to_cloud_storage("alchemy-assessment-output-001", temp_workbook_path, workbook_filename)
        workbook_file_path = f"https://storage.googleapis.com/alchemy-assessment-output-001/{workbook_filename}"

    print("6. Analysis finished. Workbook has been successfully saved at:", workbook_file_path)

    # -------------------- INTERMEDIARY SECTION TO GET METRICS FROM FILLED OUT WORKBOOK FOR FRONT END ----------------

    print("7. Beginning metrics creation...")
    # Dictionary to store the intermediate results
    control_intermediate = {}

    # Process each test procedure
    for control_family in list_of_results:
        for procedure, data in control_family.items():
            
            status = data["result"] #Getting each procedure's pass/fail
            if status == 'Partial Pass' or status == 'Error':
                status = 'Fail'

            control_name = data["name"] #Getting each procedure's 'name' (really control name)
            control = procedure.rsplit('.', 1)[0] #Getting the control ID from each procedure ID

            # If control isn't in the intermediate dictionary yet, initialize it
            if control not in control_intermediate:
                control_intermediate[control] = {'Pass': 0, 'Fail': 0, 'Name': control_name}

            # Increase the count for the current status (either Pass or Fail) for the current control
            control_intermediate[control][status] += 1

    # Final dictionary for control results
    control_results = {}

    # Determine the final status for each control
    for control, statuses in control_intermediate.items():
        control_name = statuses.pop('Name')

        if statuses['Pass'] > 0 and statuses['Fail'] == 0:
            status = 'Implemented'
        elif statuses['Pass'] > 0 and statuses['Fail'] > 0:
            status = 'Partially Implemented'
        else:
            status = 'Not Implemented'
        
        control_results[control] = {
            'Name': control_name,
            'Status': status,
            'ProcedurePassCount': statuses['Pass'],
            'ProcedureFailCount': statuses['Fail']
        }
            
    # --------------------------- METRICS ------------------------------

    # Initialize procedure-specific metric variables
    total_procedures = 0
    passed_procedures_count = 0
    failed_procedures_count = 0

    # Iterate through each control family's results at the procedure level
    for family_results in list_of_results:
        total_procedures += len(family_results)
        passed_procedures_count += sum(1 for data in family_results.values() if data["result"] == 'Pass')
    
    failed_procedures_count = total_procedures - passed_procedures_count

    #Create control-specific metric variables
    total_controls = len(control_results)
    implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Implemented')
    partially_implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Partially Implemented')
    not_implemented_controls_count = total_controls - implemented_controls_count - partially_implemented_controls_count

    # Pack metrics into a dictionary to send to the frontend

    overview_metrics = {
        'control_results': control_results,
        'total_procedures': total_procedures,
        'passed_procedures_count': passed_procedures_count,
        'failed_procedures_count': failed_procedures_count,
        'total_controls': total_controls,
        'implemented_controls_count': implemented_controls_count,
        'partially_implemented_controls_count': partially_implemented_controls_count,
        'not_implemented_controls_count': not_implemented_controls_count,
        'analyzed_file_summaries': analyzed_file_summaries,
        'number_of_files': len(analyzed_file_summaries)
    }

    print("8. Metrics creation complete...")

    # ---------------------------------- SECTION TO TEMPORARILY SAVE OUTPUTS (METRICS & TESTING WORKBOOK) FOR FRONT END ---------------------

    metrics_filename = "metrics.json"
    metrics_content = json.dumps(overview_metrics)
    
    if settings.DEBUG:
        # Save locally in development
        metrics_file_path = os.path.join(settings.STATICFILES_DIRS[0], metrics_filename)
        with open(metrics_file_path, 'w') as json_file:
            json_file.write(metrics_content)
    else:
        # Save to cloud storage in production
        temp_metrics_path = os.path.join(tempfile.mkdtemp(), metrics_filename)
        with open(temp_metrics_path, 'w') as json_file:
            json_file.write(metrics_content)
        
        save_to_cloud_storage("alchemy-assessment-output-001", temp_metrics_path, metrics_filename)
        metrics_file_path = f"https://storage.googleapis.com/alchemy-assessment-output-001/{metrics_filename}"

    print(f"9. 'metrics.json' has been successfully created at: {metrics_file_path}...")
    
    # Generate URLs for the files to let the front end access the files
    if settings.DEBUG: #If local environment - create urls for front end to access
        workbook_url = settings.STATIC_URL + os.path.basename(workbook_file_path) if workbook_file_path else None
        metrics_url = settings.STATIC_URL + os.path.basename(metrics_file_path) if metrics_file_path else None
    else: #If production - just provide the google cloud storage file path since it is a URL already
        workbook_url = workbook_file_path  # Direct URL to cloud storage
        metrics_url = metrics_file_path  # Direct URL to cloud storage

    # Construct the response data
    response_data = {
        'workbook_url': workbook_url,
        'metrics_url': metrics_url
    }

    # Return the paths of the workbook and metrics file
    return response_data
    
# Helper function to /generate_ai_assessment that takes in the worksheet, name of the organization, and the prompt template, and updates the workbook accordingly with the output of the analysis
def assess_controls_in_worksheet(worksheet, duplicate_worksheet, organization_name, prompt_template, qa_chain):

    procedure_results = {} #Initialize a dictionary to store the results of each procedure's testing

    starting_row = 2 #Start at the second row to exclude the workbook's header row

    # Loop through each row in the testing workbook and populate the prompt template with the relevant variables
    for row_index, row in enumerate(worksheet.iter_rows(min_row=starting_row, values_only=True), start=starting_row): # Assuming your data starts from the second row (skipping header)
        
        #skip the row if it is empty
        if row_index > 19:
        # if row_index == 21 or row_index == 20:

            print(f"Testing control procedure {row[2]}")

            # Getting the current control description from this row in the spreadsheet
            control_description = row[4]

            # Querying the llm to perform a similarity search with the templated query against our vector store
            question = prompt_template.format(control_description=control_description, organization=organization_name)
            result = qa_chain({"query": question})

            # print(result)

            # print(f"Result for procedure {row[2]}: {result}.")

            source_set = set()

            # Iterate through the source documents for this procedure
            for doc in result["source_documents"]:

                # Replace the full path of the Document's source with just the filename
                doc.metadata['source'] = os.path.basename(doc.metadata['source'])

                # Add the source string to the set
                source_set.add(doc.metadata['source'])
            
            # Join all metadata strings into a single string, separated by a delimiter (e.g., newline)
            all_sources = '\n'.join(source_set)

            # Set the value of the cell to the aggregated metadata string
            cell = duplicate_worksheet.cell(row=row_index, column=6, value=all_sources)
            cell.alignment = Alignment(wrap_text=True)

            # Storing the response string into separate variables by section to upload to different columns
            split_output = extract_parts(result['result'])

            #Taking the split output and populating each row's columns F through K with it
            column_index = 7

            for part in split_output:
                # Add the answer to the current column of the current row in the duplicate worksheet
                cell = duplicate_worksheet.cell(row=row_index, column=column_index, value=part)
                # Set wrap_text to True for the cell
                cell.alignment = Alignment(wrap_text=True)

                #If filling in the pass/fail Column I, then center the text in the cell horizontally and vertically
                if column_index == 9:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Move to the next column for the next loop
                column_index += 1

            procedure_id = row[2]
            procedure_name = row[3]
            procedure_result = duplicate_worksheet.cell(row=row_index, column=9).value

            procedure_results[procedure_id] = {
                "name": procedure_name,
                "result": procedure_result
            }
    
    # print("Returning testing results for " + worksheet.title + " family.")
    
    return procedure_results

# Helper function to /generate_ai_assessment that takes the output from the query to OpenAI from the similarity search, and splits it so that we can put the info into the right columns
def extract_parts(input_str):
    # Define the markers that we will use to split the text
    primary_markers = ["Analysis_Part1:", "Analysis_Part2:", "Analysis_Part3:", "Analysis_Part4:", "Analysis_Part5:"]
    secondary_markers = ["1.", "2.", "3.", "4.", "5."]

    error_message = "I'm sorry, but I can't assist with that."
    if error_message == input_str:
        return ("Error", "Error", "Error", "Error", "Error")
    
    # Function to extract parts using a given set of markers
    def extract_with_markers(markers):
        parts = []
        for i in range(len(markers)):
            start_marker = markers[i]
            try:
                end_marker = markers[i + 1]
            except IndexError:
                end_marker = None

            if end_marker:
                start_idx = input_str.find(start_marker) + len(start_marker)
                end_idx = input_str.find(end_marker)
                if start_idx == -1 or end_idx == -1:
                    return None  # Marker not found
                part = input_str[start_idx:end_idx].strip()
            else:
                start_idx = input_str.find(start_marker) + len(start_marker)
                if start_idx == -1:
                    return None  # Marker not found
                part = input_str[start_idx:].strip()
            parts.append(part)
        return parts

    # Try extracting with primary markers
    extracted_parts = extract_with_markers(primary_markers)
    if extracted_parts is not None and len(extracted_parts) == 5:
        return tuple(extracted_parts)

    # Try extracting with secondary markers
    extracted_parts = extract_with_markers(secondary_markers)
    if extracted_parts is not None and len(extracted_parts) == 5:
        return tuple(extracted_parts)

    # Handle case where neither set of markers worked
    return ('ErrorX', 'ErrorX', 'ErrorX', 'ErrorX', 'ErrorX')

@csrf_exempt
def get_assessment_status(request, task_id):

    task = AsyncResult(task_id)

    # Initialize response_data at the beginning
    response_data = {}

    try:
        #If it's finished, grab the output zip file with the metrics and the updated test workbook, and send to the front end
        if task.successful():

            current_progress = task.result.get('current', 0)
            total = task.result.get('total', 0)

            print(f"\nTask {task_id}: {task.state} with {current_progress} out of {total} procedures assessed\n")

            # Grab the file paths for the excel workbook report and the metrics json from the task result
            workbook_file_path = task.result.get('workbook_file_path', None)
            metrics_file_path = task.result.get('metrics_file_path', None)

            # Generate URLs for the files to let the front end access the files
            workbook_url = settings.STATIC_URL + os.path.basename(workbook_file_path) if workbook_file_path else None
            metrics_url = settings.STATIC_URL + os.path.basename(metrics_file_path) if metrics_file_path else None

            # Construct the response data
            response_data = {
                'state': task.state,
                'workbook_url': workbook_url,
                'metrics_url': metrics_url
            }
        
        elif task.failed():
            # Handle failed task
            response_data = {
                'state': task.state,
                'error': str(task.info),  # Error details
            }

        #If it is still ongoing, send the ongoing status and details to the front end for the progress bar
        else:

            current_progress = task.result.get('current', 0)
            total = task.result.get('total', 0)

            print(f"\nTask {task_id}: {task.state} with {current_progress} out of {total} procedures assessed\n")
            
            response_data = {
                'state': task.state,
                'details': task.info if task.info else {}
            }
        
        return JsonResponse(response_data)

    except Exception as e:
        print(f"Error in get_assessment_status: {e}")
        return JsonResponse({'error': str(e)}, status=500)

####################################### Internal Application once logged in ##############################################

# Handles showing dashboard page (for GET) as well as creating new systems before showing dashboard page again (for POST)
@login_required
def dashboard(request, system=None):

    client = request.user.client
    systems = System.objects.filter(client=client)

    if system is not None:
        system = unquote(system)

    # If the user is attempting to create a system:
    if (request.method == 'POST') & (request.POST.get("system_name") != None):

        colors = ['3E4756', '97ACCF', 'FFE1E9', 'CAAAB2', '6E7788', '5A3F46', '8D6F77', 'A2ACBD', 'CF9EAA', '986A76', '564147', 'CF9EAC', 'F3E7D0', 'BBB099', 'BEA5AB', 'BBB19B', '867D67']

        # process the data for new system. Use the client of the logged in user, and choose a random color
        system_name = request.POST.get("system_name")

        color = random.choice(colors)

        # Try to make a new system, if not possible, it's due to the system already existing
        try:
            system = System.objects.create(name=system_name, client=client, color=color)
            system.save()

        except IntegrityError:
            return render(request, "dashboard.html", {
                "message": "Email already being used."
            })

        return HttpResponseRedirect(reverse("alchemy:dashboard"))
    
    else:

        # Otherwise, if the user is just visiting the dashboard via get request, show the systems (and formulate %s for progress bar)
        if system is None:
            control_implementations_dict = {}

            for sys in systems:

                # Check that there is a statement for every ControlImplementationStatement linked to this control implementation
                empty_statement_exists = ControlImplementationStatement.objects.filter(
                    control_implementation=OuterRef('pk'),
                    statement__exact=''  # Check for empty statements
                )

                # Check that any ControlImplementationStatement is linked to this control implementation
                any_statement_exists = ControlImplementationStatement.objects.filter(
                    control_implementation=OuterRef('pk')
                )

                filtered_control_implementations = ControlImplementation.objects.filter(system=sys).annotate(
                    has_empty_statements=Exists(empty_statement_exists),
                    has_any_statements=Exists(any_statement_exists),
                    originations_count=Count('originations'),
                    statuses_count=Count('statuses'),
                ).filter(
                    responsible_role__isnull=False,  # Check if the responsible role is not null
                    originations_count__gte=1,  # Check if there is at least one ControlOrigination linked
                    statuses_count__gte=1,  # Check if there is at least one ImplementationStatus linked
                ).exclude(
                    has_empty_statements=True,  # Exclude ControlImplementations that have empty ControlImplementationStatements
                ).filter(
                    Q(has_any_statements=True) | (Q(has_any_statements=False) & ~Q(statement__exact=''))  # If ControlImplementation has no ControlImplementationStatements, it must have its statement field filled
                )

                # Count the number of ControlImplementations meeting the above criteria
                control_implementations_count = filtered_control_implementations.count()

                control_implementations_dict[sys.name] = ceil(100*(control_implementations_count/(ControlImplementation.objects.filter(system=sys).count())))

            return render(request, "internal/dashboard.html", {
                "client": client,
                "systems": systems,
                "percentages":control_implementations_dict
            })

@login_required
def delete_system(request):

    client = request.user.client

    if request.method == 'POST':
        system_name = unquote(request.POST['delete-system-button'])
        system = get_object_or_404(System, name=system_name, client=client)
        system.delete()

    return redirect('alchemy:dashboard')

@login_required
def rename_system(request):

    client = request.user.client
    

    if request.method == 'POST':
        old_system_name = unquote(request.POST['change-name-button'])
        new_system_name = unquote(request.POST['system_name'])

        system = get_object_or_404(System, name=old_system_name, client=client)
        system.name = new_system_name
        system.save(update_fields=["name"])

    return redirect('alchemy:dashboard')

@csrf_exempt
@login_required
def create_update_org(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST request required."}, status=400)

    else:
        data = json.loads(request.body)
        client_id = data.get('id', None)

        if client_id:
            client = get_object_or_404(Client, id=client_id)
            form = OrganizationForm(data, instance=client)
        else:
            form = OrganizationForm(data)

        if form.is_valid():
            client = form.save()
            response_data = {'success': True, 'client_id': client.id}
        else:
            response_data = {'success': False, 'errors': form.errors}

        return JsonResponse(response_data)
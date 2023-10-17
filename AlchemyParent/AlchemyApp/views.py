import json #for api calls
import re
import random #for generating system color for client dashboard
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse #for API calls
from django.urls import reverse #for HttpResponseRedirect(reverse)
from django.contrib.auth import authenticate, login, logout #for login/logout/register
from django.views.decorators.csrf import csrf_exempt #for API calls
from django.contrib import messages #for register error message(s)
from .models import CustomUser, Client, ControlFamily, System, ControlImplementationStatement #for interacting with database
from .forms import OrganizationForm
from django.contrib.auth.backends import ModelBackend
from django.db import IntegrityError
from django.contrib.auth.decorators import login_required #to redirect user to login route if they try to access an app page past login
from django.db.models import Count, Q, OuterRef, Exists
import os, openai
from dotenv import load_dotenv #for access to .env variables, like OPENAI API key
from urllib.parse import unquote #to decode url strings passed through urls as parameters, e.g. client
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from copy import copy
from django.conf import settings
from math import ceil
import xml.etree.ElementTree as ET
import zipfile
import tempfile
import shutil #for removing temporary directories once they are done being used

from langchain.document_loaders import PyPDFLoader #for taking relevant pdfs and loading them as langchain "document" objects
from langchain.document_loaders import Docx2txtLoader #for taking relevant word documents and loading them as langchain "document" objects
from langchain.text_splitter import CharacterTextSplitter #for taking langchain documents and splitting them into chunks (pre-processing for vector storage)
from langchain.embeddings.openai import OpenAIEmbeddings #for taking document chunks and embedding them as vectors for similarity searching
from langchain.vectorstores import FAISS #for storing the vector representations of document chunks, vectorizing the given query, and retrieving the relevant text via similarity search. Will not be long term solution
from langchain.chains import RetrievalQA #Langchain chain for distilling the retrieved document chunks into an human-like answer using an llm/chat model, like gpt-turbo-3.5
from langchain.chat_models import ChatOpenAI #Importing langchain's abstraction on top of the OpenAI API
from langchain.prompts import PromptTemplate #for creating a prompt template that will provide context to user queries/other inputs to llm

load_dotenv()

####################################### Public Application when not logged in ##############################################

# Route to render the landing page
def index(request):

    if not request.user.is_authenticated:

        return render(request, "public/index.html")
    
    else:
        return HttpResponseRedirect(reverse("alchemy:dashboard"))

def login_view(request):
    # if this is a POST, process the login data and redirect the logged in user to the overview page
    if request.method == "POST":
        # process the data in form.cleaned_data as required
        email = request.POST["email"]
        password = request.POST["password"]

        #attempt to sign user in, return User object if so
        user = authenticate(request, email=email, password=password)

        if user:
            login(request, user)
            return HttpResponseRedirect(reverse("alchemy:dashboard"))
        else:
            context = {
                'error_message': 'Invalid email or password.',
            }

            return render(request, "public/login.html", context)

    # if not a POST request, show the login page
    if not request.user.is_authenticated:
        return render(request, "public/login.html")
    else:
        return HttpResponseRedirect(reverse("alchemy:logout"))

def register_view(request):

    if request.method == 'POST':
        email = request.POST['email']
        phone_number = request.POST['phone_number']
        client = Client.objects.get(id=request.POST['client'])
        password = request.POST['password']

        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, "Email already exists.")
            return redirect('alchemy:register')

        user = CustomUser.objects.create(email=email, phone_number=phone_number, client=client)
        user.set_password(password)
        user.save()

        # Log the user in.
        backend = ModelBackend()
        user.backend = f"{backend.__module__}.{backend.__class__.__name__}"
        login(request, user)
        return redirect('alchemy:index')

    else:
        if request.user.is_authenticated:
            logout(request)
        return render(request, 'public/register.html', {
            "clients": Client.objects.all()
        })

def logout_view(request):

    if request.user.is_authenticated:
        logout(request)
    return HttpResponseRedirect(reverse("alchemy:index"))

def contact(request):

    if request.user.is_authenticated:
         return HttpResponseRedirect(reverse("alchemy:dashboard"))
    return render(request, "public/contact.html")

def assess(request):

    # render assess.html

    return render(request, "public/assess.html")

@csrf_exempt
def generate_ai_assessment(request):

    # Cannot call this API call to OpenAI unless via the file upload POSTing to backend
    if request.method != "POST":
        return JsonResponse({"error": "POST request required."}, status=400)

    # Store the OpenAI API Key for future use
    openai_api_key = os.getenv("OPENAI_API_KEY", None)
    
    # Get the relevant info from the user (files, settings)
    uploaded_files = request.FILES.getlist('data_files')
    selected_framework = request.POST['selectedFramework']
    selected_family = request.POST['selectedFamily']
    selected_model = request.POST['selectedModel']

    vectorstore = None

    for index, uploaded_file in enumerate(uploaded_files): # Loop through all uploaded files
        # If the file is too big, reject it
        if uploaded_file.multiple_chunks():
            return JsonResponse({"error": "Uploaded file is too big (%.2f MB)." % (uploaded_file.size/(1000*1000),)}, status=400)

        # Create a temporary file to save the uploaded data
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)

        # Load the temporary file using the PyPDFLoader (if a PDF ) or a DocxLoader (if a Word document)
        if uploaded_file.name.endswith('.pdf'):
            loader = PyPDFLoader(temp_file.name)
        elif uploaded_file.name.endswith('.docx'):
            loader = Docx2txtLoader(temp_file.name)
        else:
            return JsonResponse({"error": "Unsupported file type."}, status=400)

        #chunk up the loaded document. Can tweak chunking parameters for optimal performance 
        text_splitter = CharacterTextSplitter(        
            separator = "\n",
            chunk_size = 1000,
            chunk_overlap  = 200,
            length_function = len,
        )

        pages = loader.load_and_split(text_splitter)

        if index == 0:
            # Take the first file's page chunks and store them as vector embeddings within FAISS vector storage.
            vectorstore = FAISS.from_documents(pages, OpenAIEmbeddings(openai_api_key=openai_api_key))
        
        else:
            # Add every subsequent file's page chunks into the initialized vector storage
            vectorstore.add_documents(pages)

    # Instantiate the llm we'll be using to analyze the documents using the user's selected OpenAI model
    llm = ChatOpenAI(model_name=selected_model, temperature=0, openai_api_key=openai_api_key)
    
    #Instantiate the query/answer chain we'll use with the llm and PDF's vector store for answering any questions about the document.
    qa_chain = RetrievalQA.from_chain_type(llm,retriever=vectorstore.as_retriever())

    template = """{organization} has implemented security processes based on the provided documents. Please analyze the documents against the following information security test procedure: {control_description}. For this test procedure, please respond with exactly each of the 4 items, with each prefaced by their "Analysis_Part" + the number of the section, e.g. "Analysis_Part1:" (without any introduction or explanation, just your analysis):

    1. Without any introduction or explanation, the snippets of the relevant sections of text that offer evidence to support that the test requirement is implemented. Include the page number, where possible. (if it is not at all applicable, do not provide anything);
    2. An analysis of how well the document(s) meets the requirements of the given test procedure;
    3. An implementation status based on 1. and 2. of "Pass", "Fail"; and
    4. If the status was deemend "Fail" in 3., then recommendations for control remediation. If it was deemed "Pass", then do not provide anything.
    """

    # """{organization} has implemented security processes based on the provided document. How well does the provided document meet the requirements of the following information security control: '{control_description}'. Please also provide recommendations for remediation if the control is not fully met."""
    prompt_template = PromptTemplate.from_template(template)

    document_org = qa_chain({"query": "What is the name of the organization that the document is for? Reply with the name and nothing else."})['result']

    #Pull the testing workbook
    static_path = settings.STATIC_ROOT if settings.STATIC_ROOT else settings.STATICFILES_DIRS[0]
    template_path = os.path.join(static_path, 'SRC_Innovation_Documents/FedRAMP Assessment Workbook.xlsx')

    # Load workbook and select the specific testing worksheet in the workbook
    workbook = load_workbook(filename=template_path)
    
    worksheet = workbook[selected_family]

    # Create a duplicate of the original workbook for output
    duplicate_workbook = copy(workbook)
    duplicate_worksheet = duplicate_workbook[selected_family]

    # Perform analysis based on the test procedures in the controls workbook, the prompt and QA chain, and update the duplicate workbook with the output
    procedure_results = assess_controls_in_workbook(worksheet, duplicate_worksheet, document_org, prompt_template, qa_chain)

    # -------------------- SECTION TO GET METRICS FROM FILLED OUT WORKBOOK FOR FRONT END ----------------

    # Dictionary to store the intermediate results
    control_intermediate = {}

    # Process each test procedure
    for procedure, data in procedure_results.items():
        
        status = data["result"] #Getting each procedure's pass/fail
        control_name = data["name"] #Getting each procedure's 'name' (really control name)
        control = procedure.rsplit('.', 1)[0] #Getting the control ID from each procedure ID

        # If control isn't in the intermediate dictionary yet, initialize it
        if control not in control_intermediate:
            control_intermediate[control] = {'Pass': 0, 'Fail': 0, 'Name': control_name}

        # Increase the count for the current status (either Pass or Fail) for the current control
        control_intermediate[control][status] += 1

    # Final dictionary for control results
    control_results = {}

    # Determine the final status for each control
    for control, statuses in control_intermediate.items():
        control_name = statuses.pop('Name')

        if statuses['Pass'] > 0 and statuses['Fail'] == 0:
            status = 'Implemented'
        elif statuses['Pass'] > 0 and statuses['Fail'] > 0:
            status = 'Partially Implemented'
        else:
            status = 'Not Implemented'
        
        control_results[control] = {
            'Name': control_name,
            'Status': status
        }
            
    # Metrics
    total_procedures = len(procedure_results)
    passed_procedures_count = sum(1 for data in procedure_results.values() if data["result"] == 'Pass')
    failed_procedures_count = total_procedures - passed_procedures_count

    total_controls = len(control_results)
    implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Implemented')
    partially_implemented_controls_count = sum(1 for data in control_results.values() if data['Status'] == 'Partially Implemented')
    not_implemented_controls_count = total_controls - implemented_controls_count - partially_implemented_controls_count

    # Pack metrics into a dictionary to send to the frontend

    overview_metrics = {
        'control_results': control_results,
        'total_procedures': total_procedures,
        'passed_procedures_count': passed_procedures_count,
        'failed_procedures_count': failed_procedures_count,
        'total_controls': total_controls,
        'implemented_controls_count': implemented_controls_count,
        'partially_implemented_controls_count': partially_implemented_controls_count,
        'not_implemented_controls_count': not_implemented_controls_count
    }


    # ---------------------------------- SECTION TO TEMPORARILY SAVE OUTPUTS (METRICS & TESTING WORKBOOK) FOR FRONT END ---------------------

     # Create a temporary directory
    temp_dir = tempfile.mkdtemp()

    excel_file_path = os.path.join(temp_dir, "Assessment Workbook with Analysis.xlsx")

    # Save the testing workbook to a temporary file within a temporary directory
    duplicate_workbook.save(excel_file_path)

    # Save the metrics to a JSON file in the temporary directory
    json_file_path = os.path.join(temp_dir, "metrics.json")
    with open(json_file_path, 'w') as json_file:
        json.dump(overview_metrics, json_file)

    # Zip both files together
    zip_filename = "Assessment_and_Metrics.zip"
    zip_file_path = os.path.join(temp_dir, zip_filename)
    with zipfile.ZipFile(zip_file_path, 'w') as zipf:
        zipf.write(excel_file_path, os.path.basename(excel_file_path))
        zipf.write(json_file_path, os.path.basename(json_file_path))
    
    # Send the ZIP file as a response
    with open(zip_file_path, 'rb') as f:
        response = HttpResponse(f.read(), content_type="application/zip")
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    
    # Clean up the temporary directory
    shutil.rmtree(temp_dir)
    
    return response
    
    # # Read the saved workbook into memory
    # with open(temp_file_name, 'rb') as f:
    #     file_data = f.read()

    # # Remove the temporary file
    # os.remove(temp_file_name)

    # # Create an HTTP response with the file
    # response = HttpResponse(file_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename="Assessment Workbook with Analysis.xlsx"'

    # return response

# Helper function to /generate_ai_assessment that takes in the worksheet, name of the organization, and the prompt template, and updates the workbook accordingly with the output of the analysis
def assess_controls_in_workbook(worksheet, duplicate_worksheet, organization_name, prompt_template, qa_chain):

    procedure_results = {}

    starting_row = 2

    # Count rows that have data for command line percentage updates
    total_rows = sum(1 for _ in worksheet.iter_rows(values_only=True))-1

    # Loop through each row in the testing workbook and populate the prompt template with the relevant variables
    for row_index, row in enumerate(worksheet.iter_rows(min_row=starting_row, values_only=True), start=starting_row): # Assuming your data starts from the second row (skipping header)
        
        #skip the row if it is empty
        if row_index >= 21:

            # Getting the current control description from this row in the spreadsheet
            control_description = row[4]

            # Querying the llm to perform a similarity search with the templated query against our vector store
            query = prompt_template.format(control_description=control_description, organization=organization_name)
            result = qa_chain({"query": query})

            # Storing the response string into separate variables by section to upload to different columns
            split_output = extract_parts(result['result'])

            #Taking the split output and populating each row's columns G through J with it
            column_index = 8

            for part in split_output:
                # Add the answer to the current column of the current row in the duplicate worksheet
                cell = duplicate_worksheet.cell(row=row_index, column=column_index, value=part)
                # Set wrap_text to True for the cell
                cell.alignment = Alignment(wrap_text=True)

                #If filling in the pass/fail Column I, then center the text in the cell horizontally and vertically
                if column_index == 10:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Move to the next column for the next loop
                column_index += 1

            procedure_id = row[2]
            procedure_name = row[3]
            procedure_result = duplicate_worksheet.cell(row=row_index, column=10).value

            procedure_results[procedure_id] = {
                "name": procedure_name,
                "result": procedure_result
            }

            # progress bar in command line
            print(f"{round(100*((row_index-1)/(total_rows)))}% complete")
    
    return procedure_results

# Helper function to /generate_ai_assessment that takes the output from the query to OpenAI from the similarity search, and splits it so that we can put the info into the right columns
def extract_parts(input_str):
    # Define the markers that we will use to split the text
    markers = ["Analysis_Part1:", "Analysis_Part2:", "Analysis_Part3:", "Analysis_Part4:"]
    
    # Initialize an empty list to hold the parts
    parts = []
    
    # Iterate through the markers to split the text
    for i in range(len(markers)):
        start_marker = markers[i]
        try:
            end_marker = markers[i + 1]
        except IndexError:
            end_marker = None
            
        if end_marker:
            # Find the indices of the start and end markers
            start_idx = input_str.index(start_marker) + len(start_marker)
            end_idx = input_str.index(end_marker)
            
            # Extract the part in between the markers and strip whitespace
            part = input_str[start_idx:end_idx].strip()
        else:
            # We are at the last marker; extract everything after it
            start_idx = input_str.index(start_marker) + len(start_marker)
            part = input_str[start_idx:].strip()
        
        # Append the part to the list
        parts.append(part)
        
    if len(parts) != 4:
        raise ValueError("Invalid input: Expected exactly 4 parts.")
        
    return tuple(parts)

####################################### Internal Application once logged in ##############################################

# Handles showing dashboard page (for GET) as well as creating new systems before showing dashboard page again (for POST)
@login_required
def dashboard(request, system=None):

    client = request.user.client
    systems = System.objects.filter(client=client)

    if system is not None:
        system = unquote(system)

    # If the user is attempting to create a system:
    if (request.method == 'POST') & (request.POST.get("system_name") != None):

        colors = ['3E4756', '97ACCF', 'FFE1E9', 'CAAAB2', '6E7788', '5A3F46', '8D6F77', 'A2ACBD', 'CF9EAA', '986A76', '564147', 'CF9EAC', 'F3E7D0', 'BBB099', 'BEA5AB', 'BBB19B', '867D67']

        # process the data for new system. Use the client of the logged in user, and choose a random color
        system_name = request.POST.get("system_name")

        color = random.choice(colors)

        # Try to make a new system, if not possible, it's due to the system already existing
        try:
            system = System.objects.create(name=system_name, client=client, color=color)
            system.save()

        except IntegrityError:
            return render(request, "dashboard.html", {
                "message": "Email already being used."
            })

        return HttpResponseRedirect(reverse("alchemy:dashboard"))
    
    else:

        # Otherwise, if the user is just visiting the dashboard via get request, show the systems (and formulate %s for progress bar)
        if system is None:
            control_implementations_dict = {}

            for sys in systems:

                # Check that there is a statement for every ControlImplementationStatement linked to this control implementation
                empty_statement_exists = ControlImplementationStatement.objects.filter(
                    control_implementation=OuterRef('pk'),
                    statement__exact=''  # Check for empty statements
                )

                # Check that any ControlImplementationStatement is linked to this control implementation
                any_statement_exists = ControlImplementationStatement.objects.filter(
                    control_implementation=OuterRef('pk')
                )

                filtered_control_implementations = ControlImplementation.objects.filter(system=sys).annotate(
                    has_empty_statements=Exists(empty_statement_exists),
                    has_any_statements=Exists(any_statement_exists),
                    originations_count=Count('originations'),
                    statuses_count=Count('statuses'),
                ).filter(
                    responsible_role__isnull=False,  # Check if the responsible role is not null
                    originations_count__gte=1,  # Check if there is at least one ControlOrigination linked
                    statuses_count__gte=1,  # Check if there is at least one ImplementationStatus linked
                ).exclude(
                    has_empty_statements=True,  # Exclude ControlImplementations that have empty ControlImplementationStatements
                ).filter(
                    Q(has_any_statements=True) | (Q(has_any_statements=False) & ~Q(statement__exact=''))  # If ControlImplementation has no ControlImplementationStatements, it must have its statement field filled
                )

                # Count the number of ControlImplementations meeting the above criteria
                control_implementations_count = filtered_control_implementations.count()

                control_implementations_dict[sys.name] = ceil(100*(control_implementations_count/(ControlImplementation.objects.filter(system=sys).count())))

            return render(request, "internal/dashboard.html", {
                "client": client,
                "systems": systems,
                "percentages":control_implementations_dict
            })

@login_required
def delete_system(request):

    client = request.user.client

    if request.method == 'POST':
        system_name = unquote(request.POST['delete-system-button'])
        system = get_object_or_404(System, name=system_name, client=client)
        system.delete()

    return redirect('alchemy:dashboard')

@login_required
def rename_system(request):

    client = request.user.client
    

    if request.method == 'POST':
        old_system_name = unquote(request.POST['change-name-button'])
        new_system_name = unquote(request.POST['system_name'])

        system = get_object_or_404(System, name=old_system_name, client=client)
        system.name = new_system_name
        system.save(update_fields=["name"])

    return redirect('alchemy:dashboard')

@csrf_exempt
@login_required
def create_update_org(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST request required."}, status=400)

    else:
        data = json.loads(request.body)
        client_id = data.get('id', None)

        if client_id:
            client = get_object_or_404(Client, id=client_id)
            form = OrganizationForm(data, instance=client)
        else:
            form = OrganizationForm(data)

        if form.is_valid():
            client = form.save()
            response_data = {'success': True, 'client_id': client.id}
        else:
            response_data = {'success': False, 'errors': form.errors}

        return JsonResponse(response_data)